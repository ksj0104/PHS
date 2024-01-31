import os
import random
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

info_head = ["응시 문제", "최종 점수", "제출 언어", "단계1점수", "단계2점수", "단계3점수", "단계4점수", "단계5점수"]
guide_head = ["응시 문제", "종합 난이도", "1단계", "2단계", "3단계", "4단계", "5단계", "1단계 알고리즘", "2단계 알고리즘",
              "3단계 알고리즘", "4단계 알고리즘", "5단계 알고리즘", "1단계 통과 점수", "2단계 통과 점수", "3단계 통과 점수",
              "4단계 통과 점수", "5단계 통과 점수", "1단계 점수", "2단계 점수", "3단계 점수", "4단계 점수", "5단계 점수"]

# 추가조건형
problem_properties_step = ["응시 문제", "응시자 수", "전체 난이도",
                           "1단계 난이도", "2단계 난이도", "3단계 난이도", "4단계 난이도", "5단계 난이도",
                           "1단계 통과자 수", "2단계 통과자 수", "3단계 통과자 수", "4단계 통과자 수", "5단계 통과자 수",
                           "1단계 통과 비율", "2단계 통과 비율", "3단계 통과 비율", "4단계 통과 비율", "5단계 통과 비율",
                           "1단계 점수 분포", "2단계 점수 분포", "3단계 점수 분포", "4단계 점수 분포", "5단계 점수 분포",
                           "평균", "표준편차", "분산"]

used_problem_title_list = set([])
algorithm_step_passrate = {} # 알고리즘별, 난이도 별 통과 비율?
step_problem = {}

width = 1900
height = 900


selected_level_guide = None
selected_filepath = None
selected_value = [None, None, None]
selected_option = [None, None, None]

selected_problem = None
selected_problem_unsed = None

filter = {}

Tbody = []  # 전체 데이터

Curbody = []  # 필터링된 현재 데이터
cur_page = 0
cur_start = 0
cur_end = 0
total_data = 0
page_view_cnt = 23  # 소수로 해주면 좋고..

level_similarity = 1

# 그래프 그리는 도구
fig = Figure(figsize=(6, 4), dpi=80)
ax = fig.add_subplot(111)
ax1 = ax.twinx()

fig2 = Figure(figsize=(6, 4), dpi=80)
ax2 = fig2.add_subplot(111)


def show_warning():
    messagebox.showwarning("경고", "개발자 역검 결과 Excel 파일을 먼저 업로드해주세요")

def file_aggregator():
    global selected_filepath
    # aggregated_file_name = "C:/total_result.xlsx"
    aggregated_file_name = "C:/Users/ksj0104/Downloads/total_result.xlsx"



    new_book = Workbook()
    new_sheet = new_book.active

    # head 추가
    new_sheet.append(info_head)

    file_paths = filedialog.askopenfilenames(
        initialdir="C:/Users/ksj0104/Downloads",
        # initialdir="C:/",
        title="엑셀 파일 선택",
        filetypes=(("엑셀 파일", "*.xlsx"), ("모든 파일", "*.*"))
    )

    if len(file_paths) == 0:
        return

    read_cached_data = 0
    if len(file_paths) == 1 and file_paths[0] == aggregated_file_name:
        read_cached_data = 1

    for file_path in file_paths:

        df = pd.read_excel(file_path)
        df = df.fillna('-')
        df.to_excel(file_path, index=False)

        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        fourth_row = [cell.value for cell in sheet[4]]

        if '매칭 환산 점수' in fourth_row: # 역검 + 개구검 시트
            # 응시 문제	최종 점수	제출 언어	단계1점수	단계2점수	단계3점수	단계4점수	단계5점수
            for row in sheet.iter_rows(min_row=5, values_only=True):
                data = []
                for d in row:
                    data.append(d)

                indices = [index for index, value in enumerate(fourth_row) if value == '최종 점수']
                for idx in indices:
                    # idx, idx+1, idx+2, .. idx+5
                    # 문제 명, 최종 점수, 제출 언어, 단계별 ...
                    if data[idx] == '-' or data[idx+1] == '-':
                        continue
                    new_sheet.append(
                        [data[idx - 1].split('(')[0], data[idx], data[idx + 1], data[idx + 2], data[idx + 3], data[idx + 4],
                         data[idx + 5], data[idx + 6]])

        else: # 개구검 시트
            second_row = [cell.value for cell in sheet[2]]
            indices = [index for index, value in enumerate(fourth_row) if value == '최종 점수']
            for row in sheet.iter_rows(min_row=5, values_only=True):
                data = []
                for d in row:
                    data.append(d)
                # '최종 점수', '제출 언어', '단계1점수', '단계2점수', '단계3점수', '단계4점수', '단계5점수'
                for idx in indices:
                    # idx, idx+1, idx+2, .. idx+5
                    # 문제 명, 최종 점수, 제출 언어, 단계별 ...
                    if data[idx] == '-' or data[idx+1] == '-':
                        continue
                    new_sheet.append(
                        [second_row[idx - 1].split('(')[0], data[idx], data[idx + 1], data[idx + 2], data[idx + 3], data[idx + 4],
                         data[idx + 5], data[idx + 6]])

    if read_cached_data == 0:
        os.remove(aggregated_file_name)
        new_book.save(filename=aggregated_file_name)

    selected_filepath = aggregated_file_name
    show_label.config(text=f" : {selected_filepath}")
    read_excel_file()
    return

def read_excel_file():
    global filter, Tbody, Curbody
    listbox1.delete(0, tk.END)
    used_step_problem_list.delete(0, tk.END)
    unused_step_problem_list.delete(0, tk.END)
    if selected_filepath:
        try:
            workbook = openpyxl.load_workbook(selected_filepath)
            sheet = workbook.active

            Tbody = []
            Curbody = []

            for i in info_head:
                filter[i] = set([])

            for i in info_head:
                listbox1.insert(tk.END, f"{i}")

            # 선택 이벤트를 처리할 함수 연결
            listbox1.bind("<<ListboxSelect>>", select_option1)
            label.pack(pady=10)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                tmp = []
                for i in range(len(info_head)):
                    if row[i] == '-': # 점수인데 -인놈들
                        _str = '0'
                    else:
                        _str = str(row[i]).replace(" ","") # 공백 삭제
                    tmp.append(_str)
                    filter[info_head[i]].add(_str)


                used_problem_title_list.add(tmp[0])

                Tbody.append(tmp)
                Curbody.append(tmp)

            insert_data()
            show_label.config(text=f" 현재 파일 : {selected_filepath}")

        except Exception as e:
            print(f"오류: {e}")
    else:
        display_label.config(text="먼저 파일을 선택해주세요.")

def select_excel_guide():
    global selected_level_guide

    if selected_filepath is None:
        show_warning()
        return

    selected_level_guide = filedialog.askopenfilename(
        initialdir="C:/Users/ksj0104/jupyter",
        # initialdir="C:/",
        title="엑셀 파일 선택",
        filetypes=(("엑셀 파일", "*.xlsx"), ("모든 파일", "*.*"))
    )

    if selected_level_guide:
        guide_label.config(text=f"현재 파일: {selected_level_guide}")
        read_guide_file()
    else:
        guide_label.config(text="파일을 선택하지 않았습니다.")

def read_guide_file():
    global selected_level_guide, guide_head, step_problem, algorithm_list

    if selected_level_guide:
        try:
            workbook = openpyxl.load_workbook(selected_level_guide)
            sheet = workbook.worksheets[0]

            # 문제명, 종합난이도, 단계별난이도 1,2,3,4,5, 단계별알고리즘 1,2,3,4,5 단계별통과점수 1,2,3,4,5
            for row in sheet.iter_rows(min_row=2, values_only=True):
                problem_title = row[0].replace(" ","")
                step_problem[problem_title] = {}
                for i in range(1, 22, 1):
                    step_problem[problem_title][guide_head[i]] = row[i]

            # 알고리즘 난이도 별 예상 통과율
            sheet2 = workbook.worksheets[1]
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                algorithm_step_passrate[row[0]] = [0, 0, 0, 0, 0, 0]
                for i in range(1, 6, 1):
                    algorithm_step_passrate[row[0]][i] = float(row[i])

            used_step_problem_list.delete(0, tk.END)
            unused_step_problem_list.delete(0, tk.END)

            p_list_t = []
            for pt in step_problem:
                p_list_t.append(pt)

            p_list_t.sort()
            for pt in p_list_t:
                if pt in used_problem_title_list:
                    used_step_problem_list.insert(tk.END, f"{pt}")
                else:
                    unused_step_problem_list.insert(tk.END, f"{pt}")
                # unused_step_problem_list.insert(tk.END, f"{pt}")

            used_step_problem_list.bind("<<ListboxSelect>>", select_step_used)
            unused_step_problem_list.bind("<<ListboxSelect>>", select_step_unused)

        except Exception as e:
            print(f"오류: {e}")
    else:
        display_label.config(text="먼저 파일을 선택해주세요.")

def update_label(value):
    global level_similarity
    level_similarity = value
    init_descriptor()


def init_descriptor():  # 추가 조건형 분석 버튼을 눌렀을 때
    global step_problem, selected_problem, selected_problem_unsed, algorithm_step_passrate, guide_head # 추가조건형 문제 관련 정보, 알고리즘 단계별 난이도

    # guide_head = ["응시 문제", "종합 난이도", "1단계", "2단계", "3단계", "4단계", "5단계", "1단계 알고리즘", "2단계 알고리즘",
    #               "3단계 알고리즘", "4단계 알고리즘", "5단계 알고리즘", "1단계 통과 점수", "2단계 통과 점수", "3단계 통과 점수",
    #               "4단계 통과 점수", "5단계 통과 점수", "1단계 점수", "2단계 점수", "3단계 점수", "4단계 점수", "5단계 점수"]

    pass_rate_prediction = [0, 0, 0, 0, 0]
    score_variation_prediction = []

    weight_ = [[0, 0.01, 0.05, 0.10, 0.15, 0.19],  # 1단계 문제
               [0, 0.02, 0.06, 0.11, 0.17, 0.22],
               [0, 0.02, 0.08, 0.13, 0.19, 0.21],
               [0, 0.04, 0.10, 0.15, 0.20, 0.21],
               [0, 0, 0, 0, 0, 0]
               ]

    pass_count_real = [0, 0, 0, 0, 0]
    score_variation = []
    total_person = 0
    weight = 0.0

    if used_step_problem_list.curselection():
        # selected_problem
        level = step_problem[selected_problem][guide_head[1]]
        step_level = [step_problem[selected_problem][guide_head[i]] for i in range(2, 7, 1)]
        step_pass_score = []
        step_used_algorithm = []
        # 8 ~ 12 = 단계별 사용 알고리즘
        for i in range(7, 12, 1):
            _str = step_problem[selected_problem][guide_head[i]].split(',')
            step_used_algorithm.append(_str)

        for i in range(1, 5,1):
            step_level[i] = max(step_level[i-1], step_level[i])

        # 통과점수 12 ~ 16, 단계별 배점 17 ~ 21
        for i in range(12, 17, 1):
            step_pass_score.append(float(step_problem[selected_problem][guide_head[i]]) * float(step_problem[selected_problem][guide_head[i+5]]) / 100.0)

        for data in Tbody:
            if data[0] == selected_problem:
                total_person += 1
                score_variation.append(float(data[1]))
                for i in range(3, 8, 1):
                    if step_pass_score[i-3] <= float(data[i]):
                        pass_count_real[i-3] += 1

        for i in range(5):
            pass_count_real[i] /= total_person

        for i in range(5):
            _sum = 0.0
            for algo in step_used_algorithm[i]:
                _sum += algorithm_step_passrate[algo][step_level[i]]
            pass_rate_prediction[i] = _sum / len(step_used_algorithm[i]) * (1 - weight)
            weight += weight_[i][step_level[i]]

        ac = 0
        man = 100.0
        for i in range(4, -1, -1):
            p = pass_rate_prediction[i] - ac
            for j in range(int(p*30000)):
                # 점수 범위
                random_score = random.randint(int(man) - int(step_problem[selected_problem][guide_head[i+17]]), int(man))
                score_variation_prediction.append(random_score)

            ac += p
            man -= float(step_problem[selected_problem][guide_head[i+17]])

        draw_graph(pass_count_real, score_variation, score_variation_prediction, pass_rate_prediction)

    if unused_step_problem_list.curselection():
        # 대충 500명이 응시했을 때, 점수 분포도 및 평균 예측 점수, 상위 x % 위치 보기

        level = step_problem[selected_problem_unsed][guide_head[1]]
        step_level = [step_problem[selected_problem_unsed][guide_head[i]] for i in range(2, 7, 1)]
        step_pass_score = []
        step_used_algorithm = []
        # 8 ~ 12 = 단계별 사용 알고리즘
        for i in range(7, 12, 1):
            _str = step_problem[selected_problem_unsed][guide_head[i]].split(',')
            step_used_algorithm.append(_str)

        for i in range(1, 5,1):
            step_level[i] = max(step_level[i-1], step_level[i])

        # 통과점수 12 ~ 16, 단계별 배점 17 ~ 21
        for i in range(12, 17, 1):
            step_pass_score.append(float(step_problem[selected_problem_unsed][guide_head[i]]) * float(step_problem[selected_problem_unsed][guide_head[i+5]]) / 100.0)

        for i in range(5):
            _sum = 0.0
            for algo in step_used_algorithm[i]:
                _sum += algorithm_step_passrate[algo][step_level[i]]
            pass_rate_prediction[i] = _sum / len(step_used_algorithm[i]) * (1 - weight)
            weight += weight_[i][step_level[i]]


        ac = 0
        man = 100.0
        for i in range(4, -1, -1):
            p = pass_rate_prediction[i] - ac
            for j in range(int(p*30000)):
                # 점수 범위
                random_score = random.randint(int(man) - int(step_problem[selected_problem_unsed][guide_head[i+17]]), int(man))
                score_variation_prediction.append(random_score)

            ac += p
            man -= float(step_problem[selected_problem_unsed][guide_head[i+17]])


        draw_graph(None, None, score_variation_prediction, pass_rate_prediction)

    return

def draw_graph(parse, scores, prediction_scores, step_average_pass_rate):
    #그래프 그리자.
    pre_aver_score = sum(prediction_scores) / len(prediction_scores)

    if scores is not None:
        average_score = sum(scores) / len(scores)
    else:
        average_score = pre_aver_score

    ax.clear()
    ax1.clear()
    ax2.clear()

    ax.hist(prediction_scores, bins=50, color='r', edgecolor='black')

    if scores is not None:
        ax1.hist(scores, bins=20, color='b',edgecolor='black')  # 히스토그램 그리기

    if scores is not None:
        ax.axvline(x=average_score, color='b', linestyle='--', label= 'mean score')
    ax.axvline(x=pre_aver_score, color='r', linestyle='--', label= 'pmean score')
    ax.set_title('Score Distribution')
    ax.set_xlabel('score')
    ax.set_ylabel('person')
    ax.legend()
    canvas.draw()

    ax2.set_title('Pass Rate by Stage')
    if parse is not None:
        ax2.plot([1, 2, 3, 4, 5], parse, label='real data')

    ax2.plot([1, 2, 3, 4, 5], step_average_pass_rate, color='r', label='predicted data')
    ax2.legend()
    canvas2.draw()


    return

def select_step_used(event):
    global selected_problem
    if used_step_problem_list.curselection():
        selected_problem = used_step_problem_list.get(used_step_problem_list.curselection())
        step_problem_label.config(text=f"{selected_problem}")

    else:
        step_problem_label.config(text="문제 목록")


def select_step_unused(event):
    global selected_problem_unsed

    if unused_step_problem_list.curselection():
        selected_problem_unsed = unused_step_problem_list.get(unused_step_problem_list.curselection())
        step_problem_level_label.config(text=f"{selected_problem_unsed}")
    else:
        step_problem_level_label.config(text="미사용 문제 목록")

# 필터와 값 선택
def select_option1(event):
    global selected_option

    if listbox1.curselection():
        op1 = listbox1.get(listbox1.curselection())
        label.config(text=f"{op1}")
        list_option1.delete(0, tk.END)

        selected_value[0] = None
        value_label1.config(text=f"선택 없음")
        for data in filter[op1]:
            list_option1.insert(tk.END, f"{data}")
        list_option1.bind("<<ListboxSelect>>", select_value1)
        option_label1.config(text=f"{op1}")

        selected_option[0] = op1

def select_value1(event):
    global selected_value
    if list_option1.curselection():
        sv1 = list_option1.get(list_option1.curselection())
        selected_value[0] = sv1
        value_label1.config(text=f"{sv1}")


def insert_data():
    global tree, cur_start, cur_end
    tree = ttk.Treeview(root, columns=(info_head), show="headings")
    tree.pack(side="left", fill="both", expand=True)
    tree.place(x=10, y=300)
    tree.configure(height=page_view_cnt)  # 최대 30개의 행만 표시

    for head in info_head:
        tree.heading(head, text=head)
        tree.column(head, width=100)

    # 트리에 데이터 추가
    tree.delete(*tree.get_children())  # 기존 데이터 삭제
    cur_start = 0
    cur_end = min(page_view_cnt - 1, len(Curbody) - 1)
    show_table_body()

def show_table_body():
    tree.delete(*tree.get_children())  # 기존 데이터 삭제

    Curbody.sort(key=lambda x: float(x[1]), reverse=True)

    for i in range(cur_start, cur_end + 1, 1):
        tree.insert("", "end", values=(Curbody[i]))

#필터링
def filter_search():
    # 데이터 중 selected_option 가 selected_value 인 모든 열을 테이블에 그려준다.
    global cur_page, cur_start, cur_end, Curbody, selected_value, selected_option

    if selected_filepath is None:
        show_warning()
        return

    idxs = [-1, -1, -1]
    for i in range(0, 3, 1):
        if selected_option[i] != None and selected_value[i] != None:
            idxs[i] = info_head.index(selected_option[i])
        else:
            idxs[i] = -1

    Curbody = []
    for data in Tbody:
        flag = 1
        for i in range(0, 3, 1):
            if idxs[i] != -1:
                if data[idxs[i]] != selected_value[i]:
                    flag = 0
        if flag == 1:
            Curbody.append(data)
    cur_page = 0
    cur_start = 0
    cur_end = min(page_view_cnt - 1, len(Curbody) - 1)
    show_table_body()

#옵션 및 필터 초기화
def option_init():
    global selected_value, selected_option, Curbody, Tbody
    selected_value = [None, None, None]
    selected_option = [None, None, None]
    list_option1.delete(0, tk.END)
    option_label1.config(text="필터 없음")
    value_label1.config(text="선택 없음")

    Curbody = []

    for row in Tbody:
        Curbody.append(row)
    show_table_body()

## 페이지 이동 버튼
def next_page():
    global cur_page, cur_start, cur_end

    if selected_filepath is None:
        show_warning()
        return

    data_len = len(Curbody)
    cur_page += 1
    max_page_num = data_len // page_view_cnt
    cur_page = min(cur_page, max_page_num)
    cur_start = cur_page * page_view_cnt
    cur_end = cur_start + page_view_cnt - 1
    cur_end = min(cur_end, data_len - 1)
    show_table_body()
    return

def pre_page():
    global cur_page, cur_start, cur_end

    if selected_filepath is None:
        show_warning()
        return

    data_len = len(Curbody)
    cur_page -= 1
    cur_page = max(0, cur_page)
    cur_start = cur_page * page_view_cnt
    cur_end = cur_start + page_view_cnt - 1
    cur_end = min(data_len - 1, cur_end)
    show_table_body()
    return

def first_page():
    global cur_page, cur_start, cur_end

    if selected_filepath is None:
        show_warning()
        return

    data_len = len(Curbody)
    cur_page = 0
    cur_start = cur_page * page_view_cnt
    cur_end = cur_start + page_view_cnt - 1
    cur_end = min(data_len - 1, cur_end)
    show_table_body()
    return

def last_page():
    global cur_page, cur_start, cur_end

    if selected_filepath is None:
        show_warning()
        return

    data_len = len(Curbody)
    max_page_num = data_len // page_view_cnt
    cur_page = max_page_num
    cur_start = cur_page * page_view_cnt
    cur_end = cur_start + page_view_cnt - 1
    cur_end = min(data_len - 1, cur_end)
    show_table_body()
    return


root = tk.Tk()
root.title("개발자 역량 검사 엑셀 뷰어")
root.geometry(f"{width}x{height}")

label = tk.Label(root, text="엑셀 파일을 선택하세요:")
label.place(x=10, y=30)

show_label = tk.Label(root, text="다중 파일 선택 가능")
show_label.place(x=150, y=30)

# 엑셀 파일 선택 버튼
select_button = tk.Button(root, text="파일 열기", command=file_aggregator)
select_button.place(x=50, y=50)

first_button = tk.Button(root, text="맨 앞으로", command=first_page)
first_button.place(x=200, y=800)

next_button = tk.Button(root, text="다음 페이지", command=next_page)
next_button.place(x=400, y=800)

pre_button = tk.Button(root, text="이전 페이지", command=pre_page)
pre_button.place(x=300, y=800)

last_button = tk.Button(root, text="맨 뒤로", command=last_page)
last_button.place(x=500, y=800)

listbox1 = tk.Listbox(root, height=10)
listbox1.place(x=50, y=100)

val_option_y = 270

option_label1 = tk.Label(root, text="필터 없음")
option_label1.place(x=50, y=val_option_y)

list_option1 = tk.Listbox(root, height=10)
list_option1.place(x=200, y=100)

value_label1 = tk.Label(root, text="선택 없음")
value_label1.place(x=200, y=val_option_y)

check_button = tk.Button(root, text="추가조건형 분석", command=init_descriptor)
check_button.place(x=850, y=50)

upload_button = tk.Button(root, text="가이드 파일 업로드", command=select_excel_guide)
upload_button.place(x=1000, y=50)

guide_label = tk.Label(root, text="현재 불러온 파일:")
guide_label.place(x=1000, y=30)

step_problem_label = tk.Label(root, text="응시된 문제 목록")
step_problem_label.place(x=850, y=val_option_y)

used_step_problem_list = tk.Listbox(root, height=10)
used_step_problem_list.place(x=850, y=100)

step_problem_level_label = tk.Label(root, text="미사용 문제 목록")
step_problem_level_label.place(x=1000, y=val_option_y)

unused_step_problem_list = tk.Listbox(root, height=10)
unused_step_problem_list.place(x=1000, y=100)

# 필터 버튼
filter_button = tk.Button(root, text="필터", command=filter_search)
filter_button.place(x=250, y=50)

init_button = tk.Button(root, text="필터 초기화", command=option_init)
init_button.place(x=300, y=50)

display_label = tk.Label(root, text="", justify="left", anchor="w", font=("Courier", 10))
display_label.place(x=10, y=10)


scale = tk.Scale(root, from_=1, to=5, orient="horizontal", length=200, command=update_label)
scale.place(x=1150, y=50)

graph_canvas = tk.Canvas(root, width=300, height=300)
graph_canvas.place(x=850, y=300)

graph_canvas2 = tk.Canvas(root, width=300, height=300)
graph_canvas2.place(x=1350, y=300)

canvas = FigureCanvasTkAgg(fig, master=graph_canvas)
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

canvas2 = FigureCanvasTkAgg(fig2, master=graph_canvas2)
canvas2.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

root.mainloop()
